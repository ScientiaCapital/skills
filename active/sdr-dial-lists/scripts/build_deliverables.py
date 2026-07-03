#!/usr/bin/env python3
import json, html, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
DATE="2026-06-18"
OUT="/sessions/great-zen-curie/mnt/outputs"
FOLDER="/sessions/great-zen-curie/mnt/epiphan-mcp-guide"
ARCH=f"{FOLDER}/sdr-dial-archive"
os.makedirs(ARCH,exist_ok=True)
q=json.load(open(f"{OUT}/data/queues_final.json"))

def stats(rows):
    from collections import Counter
    return {
      "n":len(rows),
      "us":sum(1 for r in rows if r['country']=='US'),
      "ca":sum(1 for r in rows if r['country']=='CA'),
      "he":sum(1 for r in rows if r['vertical']=='Higher Ed'),
      "le":sum(1 for r in rows if r['vertical']=='Live Events'),
      "atl":sum(1 for r in rows if r['power']=='ATL'),
      "btl":sum(1 for r in rows if r['power']=='BTL'),
      "owned":sum(1 for r in rows if r['source']=='owned'),
      "gf":sum(1 for r in rows if r['source']=='greenfield'),
    }
se,sv=stats(q['edgar']),stats(q['vasil'])

# ---------------- HTML ----------------
def hrows(rows):
    out=[]
    for i,r in enumerate(rows,1):
        link=f"https://app.hubspot.com/contacts/21530819/contact/{r['cid']}"
        powcls="atl" if r['power']=="ATL" else "btl"
        srccls="owned" if r['source']=="owned" else "gf"
        note_html=html.escape(r['note']).replace("\n","<br>")
        out.append(f"""<tr class="row">
<td class="num">{i}</td>
<td><a href="{link}" target="_blank"><b>{html.escape(r['name'])}</b></a><br><span class="t">{html.escape(r['title'])}</span></td>
<td>{html.escape(r['company'])}<br><span class="t">{html.escape(r['state'])} · {r['country']}</span></td>
<td><span class="v">{r['vertical']}</span> <span class="icp">ICP {90 if r['vertical']=='Higher Ed' else 85}</span></td>
<td><span class="pw {powcls}">{r['power']}</span></td>
<td><span class="src {srccls}">{'OWNED' if r['source']=='owned' else 'GREENFIELD'}</span></td>
<td class="ph"><a href="tel:{html.escape(r['phone'])}">{html.escape(r['phone'])}</a><br><span class="t">{html.escape(r['email'])}</span></td>
<td class="note"><details><summary>note</summary><pre>{note_html}</pre></details></td>
</tr>""")
    return "\n".join(out)

def tab(name,rows,s):
    return f"""<div class="tab" id="{name}">
<div class="sb">
<span>{name.title()} · {s['n']} calls</span>
<span>US {s['us']} · CA {s['ca']}</span>
<span>HE {s['he']} · Live Events {s['le']}</span>
<span>ATL {s['atl']} · BTL {s['btl']}</span>
<span>Owned {s['owned']} · Greenfield {s['gf']}</span>
</div>
<table>
<thead><tr><th>#</th><th>Name / Title</th><th>Company</th><th>Vertical</th><th>Power</th><th>Source</th><th>Phone / Email</th><th>Spec-accurate note</th></tr></thead>
<tbody>
{hrows(rows)}
</tbody></table></div>"""

HTML=f"""<!DOCTYPE html><html><head><meta charset="utf-8"><title>SDR Dial Lists {DATE}</title>
<style>
body{{font-family:Arial,Helvetica,sans-serif;margin:0;background:#f4f5f7;color:#172b4d}}
header{{background:#172b4d;color:#fff;padding:18px 26px}}
header h1{{margin:0;font-size:20px}} header .sub{{color:#b3bac5;font-size:13px;margin-top:4px}}
.tabs{{display:flex;gap:6px;padding:12px 26px 0}}
.tabs button{{background:#dfe1e6;border:0;padding:10px 22px;font-size:14px;font-weight:bold;cursor:pointer;border-radius:6px 6px 0 0;color:#42526e}}
.tabs button.on{{background:#fff;color:#FF7A59;border-bottom:3px solid #FF7A59}}
.tab{{display:none;background:#fff;margin:0 26px 26px;padding:0 0 18px;border-radius:0 8px 8px 8px;box-shadow:0 1px 3px rgba(9,30,66,.2)}}
.tab.on{{display:block}}
.sb{{display:flex;gap:24px;flex-wrap:wrap;padding:14px 20px;background:#fafbfc;border-bottom:1px solid #ebecf0;font-size:13px;font-weight:bold;color:#42526e}}
table{{width:100%;border-collapse:collapse;font-size:13px}}
th{{background:#FF7A59;color:#fff;text-align:left;padding:9px 10px;position:sticky;top:0;font-size:12px}}
td{{padding:9px 10px;border-bottom:1px solid #ebecf0;vertical-align:top}}
.row:hover{{background:#f4f5f7}}
.num{{color:#7a869a;font-weight:bold}}
.t{{color:#7a869a;font-size:11px}}
a{{color:#0052cc;text-decoration:none}} a:hover{{text-decoration:underline}}
.v{{font-weight:bold}} .icp{{color:#7a869a;font-size:11px}}
.pw{{padding:2px 8px;border-radius:10px;font-size:11px;font-weight:bold;color:#fff}}
.pw.atl{{background:#00875a}} .pw.btl{{background:#5243aa}}
.src{{padding:2px 8px;border-radius:10px;font-size:11px;font-weight:bold}}
.src.owned{{background:#deebff;color:#0747a6}} .src.gf{{background:#fffae6;color:#974f0c}}
.note pre{{white-space:pre-wrap;font-family:Consolas,Monaco,monospace;font-size:11px;background:#fafbfc;padding:10px;border:1px solid #ebecf0;border-radius:6px;max-width:520px}}
.note summary{{cursor:pointer;color:#0052cc;font-weight:bold}}
.foot{{padding:10px 26px;color:#7a869a;font-size:12px}}
</style></head><body>
<header><h1>Epiphan SDR Dial Lists — {DATE}</h1>
<div class="sub">USA+Canada · Higher Ed + Live Events · rolling 90-day active · synced as Nooks call tasks (due 9 AM EST) · specs verified via Epiphan AI</div></header>
<div class="tabs">
<button class="on" onclick="show('edgar',this)">Edgar ({se['n']})</button>
<button onclick="show('vasil',this)">Vasil ({sv['n']})</button>
</div>
{tab('edgar',q['edgar'],se)}
{tab('vasil',q['vasil'],sv)}
<div class="foot">Discovery + talking points per study-app killer-questions + epiphan-call-playbook Verified Spec Bank. Book: meetings.hubspot.com/timkipper/discovery-call · Study: study-app-flax.vercel.app/killer-questions</div>
<script>
function show(id,btn){{document.querySelectorAll('.tab').forEach(t=>t.classList.remove('on'));document.querySelectorAll('.tabs button').forEach(b=>b.classList.remove('on'));document.getElementById(id).classList.add('on');btn.classList.add('on');}}
document.getElementById('edgar').classList.add('on');
</script></body></html>"""
open(f"{OUT}/sdr_dial_lists_{DATE}.html","w").write(HTML)
print("HTML written")

# ---------------- XLSX ----------------
HDR=PatternFill("solid",start_color="FF7A59")
HF=Font(name="Arial",bold=True,color="FFFFFF",size=11)
AF=Font(name="Arial",size=10)
BF=Font(name="Arial",bold=True,size=10)
thin=Side(style="thin",color="D9D9D9")
BORD=Border(left=thin,right=thin,top=thin,bottom=thin)
WRAP=Alignment(wrap_text=True,vertical="top")
TOP=Alignment(vertical="top")

wb=Workbook()
# Summary
ws=wb.active; ws.title="Summary"
ws["A1"]="Epiphan SDR Dial Lists"; ws["A1"].font=Font(name="Arial",bold=True,size=16)
ws["A2"]=f"{DATE} · USA+Canada · Higher Ed + Live Events · 90-day active · Nooks call tasks due 9 AM EST"; ws["A2"].font=Font(name="Arial",size=10,italic=True)
hdrs=["Rep","Total","US","CA","Higher Ed","Live Events","ATL","BTL","Owned","Greenfield"]
for j,h in enumerate(hdrs,1):
    c=ws.cell(row=4,column=j,value=h); c.fill=HDR; c.font=HF; c.border=BORD
def srow(ws,r,name,s):
    vals=[name,s['n'],s['us'],s['ca'],s['he'],s['le'],s['atl'],s['btl'],s['owned'],s['gf']]
    for j,v in enumerate(vals,1):
        c=ws.cell(row=r,column=j,value=v); c.font=BF if j==1 else AF; c.border=BORD
srow(ws,5,"Edgar Marroquin",se)
srow(ws,6,"Vasil Ivanov (ramping)",sv)
# team totals via formula
ws.cell(row=7,column=1,value="TEAM").font=BF
for j in range(2,11):
    col=chr(64+j); ws.cell(row=7,column=j,value=f"=SUM({col}5:{col}6)").font=BF; ws.cell(row=7,column=j).border=BORD
ws.cell(row=7,column=1).border=BORD
for col,w in zip("ABCDEFGHIJ",[22,8,6,6,11,12,7,7,8,11]): ws.column_dimensions[col].width=w
ws["A9"]="Notes:"; ws["A9"].font=BF
ws["A10"]="Edgar = 28 owned (HE+Live Events, US) + 22 US greenfield fill. 23 net-new Nooks call tasks created today; 27 owned carry prior open call tasks."
ws["A11"]="Vasil (ramping, all-greenfield) = 50 net-new Nooks call tasks, ATL-first, US-weighted (45 US / 5 CA)."
ws["A12"]="Scope lock: HE + Live Events only. Out-of-scope owned (broadcast/PBS/film: Olelo, WMHT, WKAR, ILM) dropped from Edgar."
for r in (10,11,12): ws.cell(row=r,column=1).font=AF

# per-rep sheets
def sheet(name,rows,s):
    w=wb.create_sheet(name)
    cols=["#","Name","Title","Company","State","Country","Vertical","ICP","Power","Source","Phone","Email","Multithread","Last Activity","HubSpot Link"]
    for j,h in enumerate(cols,1):
        c=w.cell(row=1,column=j,value=h); c.fill=HDR; c.font=HF; c.border=BORD; c.alignment=Alignment(wrap_text=True,vertical="center")
    import re
    for i,r in enumerate(rows,1):
        # extract multithread line from note
        mt=""
        m=re.search(r"--- MULTITHREAD ---\n(.+?)\n\n",r['note'],re.S)
        if m: mt=m.group(1).strip()
        link=f"https://app.hubspot.com/contacts/21530819/contact/{r['cid']}"
        vals=[i,r['name'],r['title'],r['company'],r['state'],r['country'],r['vertical'],
              90 if r['vertical']=='Higher Ed' else 85,r['power'],
              'OWNED' if r['source']=='owned' else 'GREENFIELD',r['phone'],r['email'],mt,r['last_act'],link]
        for j,v in enumerate(vals,1):
            c=w.cell(row=i+1,column=j,value=v); c.font=AF; c.border=BORD
            c.alignment=WRAP if j in (3,4,13) else TOP
        w.cell(row=i+1,column=15).hyperlink=link; w.cell(row=i+1,column=15).font=Font(name="Arial",size=10,color="0052CC")
    # stats row
    sr=len(rows)+2
    w.cell(row=sr,column=2,value=f"{s['n']} calls | US {s['us']} CA {s['ca']} | HE {s['he']} LE {s['le']} | ATL {s['atl']} BTL {s['btl']} | Owned {s['owned']} GF {s['gf']}").font=BF
    for col,wd in zip("ABCDEFGHIJKLMNO",[4,20,30,26,12,8,11,5,7,11,16,28,52,12,20]): w.column_dimensions[col].width=wd
    w.freeze_panes="A2"
sheet("Edgar",q['edgar'],se)
sheet("Vasil",q['vasil'],sv)

path=f"{OUT}/sdr_dial_lists_{DATE}.xlsx"
wb.save(path)
print("XLSX written:",path)
